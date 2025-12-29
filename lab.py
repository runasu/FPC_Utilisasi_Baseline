#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FPC Utilization capture script (Baseline + Enhancements, Windows-safe, DEBUG)
=============================================================================
Changelog (24 Dec 2025):
- FIX: Added missing helper `write_hardware_row_simple()` so hardware rows are written correctly.
- FIX: `get_area_pop_from_node()` regex now captures the leading token before '-' or space.
- FIX: FPC slot extraction regex for interfaces (et|xe|ge) was malformed; corrected.
- Minor: Safer handling for regexes, text wrapping, and IPv4 parsing.
"""
import datetime, os, re, sys, time, logging
from typing import Dict, Any, List

# --- Dependencies ---
try:
    import paramiko
except Exception:
    sys.stderr.write("Missing dependency: paramiko. Install: pip install paramiko\n"); sys.exit(1)
try:
    from xml.dom import minidom
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.drawing.image import Image
except Exception:
    sys.stderr.write("Missing dependency: openpyxl or xml.dom. Install: pip install openpyxl\n"); sys.exit(1)

# --- constants (baseline) ---
SSH_PORT = 21112
BANNER_TIMEOUT = 180
INITIAL_TEST_RETRIES = 5
INITIAL_TEST_RETRY_DELAY = 10
MAIN_SHEET = 'Utilisasi FPC'
UTIL_SHEET = 'Utilisasi Port'
ALARM_SHEET = 'Alarm Status'
HARDWARE_SHEET = 'Hardware Inventory'
SYSTEM_SHEET = 'System Performance'
DASHBOARD_SHEET = 'Dashboard Summary'
TOP_N = 5

# --- globals ---
folder_daily_global = None
folder_monthly_global = None
capture_time_global = None
debug_folder_global = None
logging_silent = False
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(SCRIPT_DIR, 'logo_lab.png')

# --- styles/colors ---
NAVY = '1F4E79'; BLUE = '4F81BD'; ORANGE = 'E67E22'; RED = 'E74C3C'; GREEN = '27AE60'; PURPLE = '8E44AD'
WHITE = 'FFFFFF'; ZEBRA_A= 'FFFFFF'; ZEBRA_B= 'F8F9FA'; LIGHT= 'F1F2F6'
THIN_BORDER = Border(left=Side(style='thin', color='D3D3D3'), right=Side(style='thin', color='D3D3D3'), top=Side(style='thin', color='D3D3D3'), bottom=Side(style='thin', color='D3D3D3'))

def fill(color): return PatternFill('solid', fgColor=color)

# --- Logging ---
logger = logging.getLogger("FPCBaselinePlus")
logger.setLevel(logging.DEBUG)

# --- helpers ---
def get_area_pop_from_node(node_name: str) -> str:
    """Return leading token (letters/numbers) before first '-' or whitespace, e.g., 'JKT' from 'JKT-DCI-INET...'."""
    s = str(node_name or '').strip()
    # Capture from beginning up to '-' or whitespace
    m = re.match(r'^([A-Za-z0-9]+)', s)
    return m.group(1).upper() if m else '-'

def _banner_start(nodes_count: int) -> None:
    tz = get_indonesia_timezone()
    line = "="*72
    print(f"\n{line}\nTELKOM+ FPC Utilization Collector (Baseline+ DEBUG)\nStart: {capture_time_global.strftime('%d %B %Y, %H:%M:%S')} {tz}\nNodes: {nodes_count}\n{line}")

def _banner_end(total_secs: float) -> None:
    line = "="*72
    print(f"{line}\nCompleted in {total_secs:.1f}s\nOutput folder: {folder_monthly_global}\n{line}\n")

def print_status(level, message, node_name=None):
    ts = datetime.datetime.now().strftime('%H:%M:%S')
    tag = f"[{level.upper():>5}]"
    who = f"[{node_name}]" if node_name else ""
    msg = f"{ts} {tag} {who} {message}"
    print(msg)
    try:
        lvl = level.lower()
        if lvl == 'debug': logger.debug(f"{who} {message}")
        elif lvl == 'error': logger.error(f"{who} {message}")
        elif lvl in ('warn','warning'): logger.warning(f"{who} {message}")
        else: logger.info(f"{who} {message}")
    except Exception:
        pass

def get_desktop_path():
    if os.name == 'nt':
        try:
            from ctypes import windll, create_unicode_buffer
            buf = create_unicode_buffer(260)
            if windll.shell32.SHGetFolderPathW(None, 0x0010, None, 0, buf) == 0:
                return buf.value
        except Exception:
            pass
    return os.path.join(os.path.expanduser('~'), 'Desktop')

def get_indonesia_timezone():
    try:
        off = time.timezone if time.daylight == 0 else time.altzone
        h = -off / 3600
        if abs(h-7)<0.5: return "WIB"
        if abs(h-8)<0.5: return "WITA"
        if abs(h-9)<0.5: return "WIT"
        return "WIB"
    except Exception:
        return "WIB"

def setup_debug_folder():
    global debug_folder_global
    if not debug_folder_global: return
    for d in ('Debug Logs','Debug XML','Temp Files'):
        try: os.makedirs(os.path.join(debug_folder_global, d), exist_ok=True)
        except Exception: pass

def get_debug_log_path(filename):
    base = folder_daily_global or '.'
    if not debug_folder_global: return os.path.join(base, filename)
    if filename.endswith('.xml'): return os.path.join(debug_folder_global, 'Debug XML', filename)
    if filename.endswith('.log'): return os.path.join(debug_folder_global, 'Debug Logs', filename)
    return os.path.join(debug_folder_global, 'Temp Files', filename)

def save_log(path, content):
    try:
        with open(path, 'w', encoding='utf-8', errors='ignore') as f: f.write(content or '')
    except Exception:
        try:
            with open(path, 'wb') as f: f.write((content or '').encode('utf-8', errors='ignore'))
        except Exception: pass

def append_error_log(path, msg):
    try:
        with open(path, 'a', encoding='utf-8', errors='ignore') as f: f.write(msg + '\n')
    except Exception: pass

_progress_start = time.monotonic()

def print_progress(current, total, node_name, operation="Processing"):
    total = total or 1
    per = (current/total)*100
    bar_len=26; filled=int((current/total)*bar_len)
    bar = "#"*filled + "."*(bar_len-filled)
    eta = (time.monotonic()-_progress_start) / max(1,current) * (total-current)
    msg = f"[{current:3d}/{total}] {operation}: {node_name:<35} [{bar}] {per:5.1f}% \n ETA ~{eta:4.1f}s"
    print(msg); logger.debug(msg)

# ---------- data row counter ----------
def _count_data_rows(ws, start_row=6, must_have_cols=(2,)):
    count = 0
    for r in range(start_row, ws.max_row + 1):
        if any(str(ws.cell(row=r, column=c).value or '').strip() for c in must_have_cols):
            count += 1
    logger.debug(f"_count_data_rows sheet={ws.title} rows={count}")
    return count

# ---------- Excel styling ----------
def style_headers(ws, header_row, bg_color):
    max_col = ws.max_column
    for c in range(1, max_col+1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = Font(name='Calibri', bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = fill(bg_color)
        cell.border = THIN_BORDER
    ws.row_dimensions[header_row].height = 24
    logger.debug(f"style_headers applied for {ws.title} row={header_row} cols={max_col}")

def style_data_rows(ws, start_row=6, wrap_cols=None):
    max_row = ws.max_row; max_col = ws.max_column
    wrap_cols = set(wrap_cols or [])
    for r in range(start_row, max_row+1):
        band = fill(ZEBRA_A if (r - start_row) % 2 == 0 else ZEBRA_B)
        for c in range(1, max_col+1):
            cell = ws.cell(row=r, column=c)
            col_letter = cell.column_letter
            want_wrap = (col_letter in wrap_cols)
            cell.font = Font(name='Calibri')
            cell.alignment = Alignment(vertical='center', wrap_text=want_wrap)
            cell.fill = band
            cell.border = THIN_BORDER
    logger.debug(f"style_data_rows applied for {ws.title} rows {start_row}-{max_row}")

def set_column_widths(ws, widths):
    for col, w in widths.items():
        try: ws.column_dimensions[col].width = w
        except Exception: pass
    logger.debug(f"set_column_widths {ws.title} widths={widths}")

def apply_table_style_like(ws, header_row=5):
    if ws.max_row > header_row:
        ws.auto_filter.ref = f"A{header_row}:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        ws.freeze_panes = f"A{header_row+1}"
    logger.debug(f"apply_table_style_like {ws.title} header_row={header_row}")

# ---------- Dynamic column resizing ----------
def dynamic_resize_columns(ws, header_row=5, start_row=6, min_w=8, max_w=46):
    try:
        max_col = ws.max_column
        widths = {}
        for c in range(1, max_col+1):
            max_len = 0
            for r in range(header_row, ws.max_row+1):
                val = ws.cell(row=r, column=c).value
                if val is None: continue
                s = str(val).replace('\n',' ')
                max_len = max(max_len, len(s))
            width = min(max_w, max(min_w, int(max_len*0.9)+2))
            widths[ws.cell(row=header_row, column=c).column_letter] = width
        set_column_widths(ws, widths)
        logger.debug(f"dynamic_resize_columns {ws.title} computed widths={widths}")
    except Exception as e:
        print_status('WARN', f'dynamic_resize_columns failed: {e}')

# ---------- Footer summaries ----------
def add_sheet_footer_summary(ws, title, bullets, start_pad=2):
    start = ws.max_row + start_pad
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=min(ws.max_column, 8))
    tcell = ws.cell(row=start, column=1, value=title)
    tcell.font = Font(name='Calibri', bold=True, size=11); tcell.alignment = Alignment(horizontal='left'); tcell.border = THIN_BORDER
    row = start + 1
    for b in bullets:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=min(ws.max_column, 8))
        cell = ws.cell(row=row, column=1, value=f"• {b}")
        cell.font = Font(name='Calibri'); cell.alignment = Alignment(horizontal='left'); cell.border = THIN_BORDER
        row += 1
    logger.debug(f"add_sheet_footer_summary {ws.title} title={title} bullets={bullets}")

# ---------- Logo helper ----------
def add_dashboard_logo(ws, logo_path=LOGO_PATH):
    try:
        if not os.path.exists(logo_path):
            print_status('WARN', f'Logo file not found: {logo_path}')
            return
        img = Image(logo_path)
        img.height = 48
        ws.add_image(img, 'A1')
        ws.column_dimensions['A'].width = 20
        ws.row_dimensions[1].height = 36
        logger.debug("Logo added to dashboard")
    except Exception as e:
        print_status('WARN', f'add_dashboard_logo failed: {e}')

# ---------- Workbook builder ----------
def workbook_create(path):
    wb = Workbook(); tz = get_indonesia_timezone()
    # DASHBOARD
    ws_dash = wb.active; ws_dash.title = DASHBOARD_SHEET; ws_dash.sheet_properties.tabColor = NAVY
    add_dashboard_logo(ws_dash, LOGO_PATH)
    ws_dash.merge_cells('B1:H1'); ws_dash['B1'] = 'NETWORK INFRASTRUCTURE MONITORING DASHBOARD'
    ws_dash['B1'].font = Font(name='Calibri', bold=True, size=16, color=WHITE)
    ws_dash['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_dash['B1'].fill = fill(NAVY)
    ws_dash.merge_cells('B2:H2'); ws_dash['B2'] = f'FPC Utilization Report - {capture_time_global.strftime("%d %B %Y, %H:%M")} {tz}'
    ws_dash['B2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws_dash['B2'].font=Font(name='Calibri'); ws_dash['B2'].fill = fill(LIGHT)
    ws_dash.merge_cells('B4:H4'); ws_dash['B4'] = 'NETWORK OVERVIEW'
    ws_dash['B4'].font = Font(name='Calibri', bold=True, color=WHITE); ws_dash['B4'].alignment = Alignment(horizontal='center')
    ws_dash['B4'].fill = fill(BLUE)
    for c, text in zip(('B5','C5','D5'), ('Metric','Count','Status')):
        ws_dash[c] = text
        ws_dash[c].font = Font(name='Calibri', bold=True, color=WHITE)
        ws_dash[c].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws_dash[c].fill = fill(BLUE); ws_dash[c].border = THIN_BORDER
    for r, txt in zip(('B6','B7','B8','B9'), ('Total Nodes','Active Interfaces','Hardware Components','System Alarms')):
        ws_dash[r] = txt; ws_dash[r].font = Font(name='Calibri')
        ws_dash[r].alignment = Alignment(horizontal='left', indent=2, wrap_text=True)
        ws_dash[r].fill = fill(ZEBRA_B); ws_dash[r].border = THIN_BORDER
    set_column_widths(ws_dash, {'A': 20, 'B': 28, 'C': 12, 'D': 18, 'E': 14, 'F': 14, 'G': 14, 'H': 14})

    # Sheets
    ws_main = wb.create_sheet(MAIN_SHEET); ws_main.sheet_properties.tabColor = GREEN
    ws_util = wb.create_sheet(UTIL_SHEET); ws_util.sheet_properties.tabColor = ORANGE
    ws_alarm = wb.create_sheet(ALARM_SHEET); ws_alarm.sheet_properties.tabColor = RED
    ws_hw = wb.create_sheet(HARDWARE_SHEET); ws_hw.sheet_properties.tabColor = PURPLE
    ws_sys = wb.create_sheet(SYSTEM_SHEET)

    def _hdr(ws, title, sub1, sub2):
        ws.merge_cells('A1:Z1'); ws['A1'] = title
        ws.merge_cells('A2:Z2'); ws['A2'] = sub1
        ws.merge_cells('A3:Z3'); ws['A3'] = sub2
        ws['A1'].font = Font(name='Calibri', bold=True, size=13)
        ws['A2'].font = Font(name='Calibri', size=11)
        ws['A3'].font = Font(name='Calibri', size=10, color="555555")
        ws['A1'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws['A2'].alignment = Alignment(horizontal='left', wrap_text=True)
        ws['A3'].alignment = Alignment(horizontal='left', wrap_text=True)
        logger.debug(f"Headers set for {ws.title}")

    _hdr(ws_main, 'NETWORK INFRASTRUCTURE MONITORING SYSTEM - FPC UTILIZATION',
         'FPC Utilization Analysis Report',
         f'Report Period: {capture_time_global.strftime("%d %B %Y, %H:%M")} {tz}')
    set_column_widths(ws_main, {'A':6,'B':28,'C':14,'D':46,'E':18,'F':22,'G':16,'H':22,'I':16,'J':16})
    headers_main = ['No.','Node Name','Divre','Interface Description','Interface ID','Module Type','Port Capacity','Current Traffic','Utilization (%)','Status']
    for i, h in enumerate(headers_main, 1): ws_main.cell(row=5, column=i, value=h)
    style_headers(ws_main, 5, NAVY); apply_table_style_like(ws_main, 5)

    _hdr(ws_util, 'NETWORK INFRASTRUCTURE MONITORING SYSTEM - PORT UTILIZATION',
         'Detailed Port Utilization Monitoring Report',
         f'Report Period: {capture_time_global.strftime("%d %B %Y, %H:%M")} {tz}')
    set_column_widths(ws_util, {'A':6,'B':30,'C':14,'D':22,'E':32,'F':16,'G':36,'H':18,'I':16,'J':46,'K':14,'L':30,'M':16})
    headers_util = ['No.','Node Name','Divre','Interface ID','Module Description','Port Capacity','Last Flapped','SFP Status','Configuration','Interface Description','Status','Flap Alert','Alert Up/Down']
    for i, h in enumerate(headers_util, 1): ws_util.cell(row=5, column=i, value=h)
    style_headers(ws_util, 5, NAVY); apply_table_style_like(ws_util, 5)

    _hdr(ws_alarm, 'NETWORK INFRASTRUCTURE MONITORING SYSTEM - ALARM STATUS',
         'Network Alarm Status Monitoring Report',
         f'Report Period: {capture_time_global.strftime("%d %B %Y, %H:%M")} {tz}')
    set_column_widths(ws_alarm, {'A':6,'B':30,'C':14,'D':30,'E':18,'F':54,'G':16,'H':16})
    headers_alarm = ['No.','Node Name','Divre','Alarm Time','Alarm Type','Alarm Description','Severity Level','Current Status']
    for i, h in enumerate(headers_alarm, 1): ws_alarm.cell(row=5, column=i, value=h)
    style_headers(ws_alarm, 5, NAVY); apply_table_style_like(ws_alarm, 5)

    _hdr(ws_hw, 'NETWORK INFRASTRUCTURE MONITORING SYSTEM - HARDWARE INVENTORY',
         'Hardware Inventory Monitoring Report',
         f'Report Period: {capture_time_global.strftime("%d %B %Y, %H:%M")} {tz}')
    set_column_widths(ws_hw, {'A':6,'B':30,'C':14,'D':20,'E':30,'F':24,'G':24,'H':50,'I':16,'J':18,'K':50})
    headers_hw = ['No.','Node Name','Divre','Component Type','Slot/Position','Part Number','Serial Number','Model/Description','Version','Operational Status','Remarks']
    for i, h in enumerate(headers_hw, 1): ws_hw.cell(row=5, column=i, value=h)
    style_headers(ws_hw, 5, NAVY); apply_table_style_like(ws_hw, 5)

    _hdr(ws_sys, 'NETWORK INFRASTRUCTURE MONITORING SYSTEM - SYSTEM PERFORMANCE',
         'System Performance Monitoring Report',
         f'Report Period: {capture_time_global.strftime("%d %B %Y, %H:%M")} {tz}')
    ws_sys.merge_cells('H4:I4'); ws_sys['H4'] = 'Memory Space'
    ws_sys.merge_cells('J4:K4'); ws_sys['J4'] = 'CPU Used'
    ws_sys.merge_cells('L4:P4'); ws_sys['L4'] = 'Hard Disk Space'
    ws_sys.merge_cells('Q4:Q4'); ws_sys['Q4'] = 'Temperature'
    for cell_addr in ('H4','J4','L4','Q4'):
        cell = ws_sys[cell_addr]
        cell.font = Font(name='Calibri', bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = fill(NAVY); cell.border = THIN_BORDER
    ws_sys.row_dimensions[4].height = 24
    set_column_widths(ws_sys, {'A':6,'B':16,'C':38,'D':22,'E':16,'F':30,'G':14,'H':14,'I':32,'J':14,'K':32,'L':18,'M':18,'N':18,'O':14,'P':30,'Q':14})
    headers_sys = ['No','Area Pop','Host Name','Loopback Address','Status Node','Current SW','Platform',
                   'Util (%)','Recommendation','Usage(%)','Recommendation',
                   'Total Space (Mbyte)','Used Space (Mbyte)','Free Space (Mbyte)','Util (%)','Recommendation','Router (°C)']
    for i, h in enumerate(headers_sys, 1): ws_sys.cell(row=5, column=i, value=h)
    style_headers(ws_sys, 5, NAVY); apply_table_style_like(ws_sys, 5)

    wb.active = wb[DASHBOARD_SHEET]
    wb.save(path); wb.close()
    logger.debug(f"Workbook created at {path}")

# ---------- write row helpers ----------
def _status_color(util_fraction: float) -> str:
    if util_fraction >= 0.80: return 'Red'
    if util_fraction >= 0.60: return 'Yellow'
    return 'Green'

# NEW: missing helper added

def write_hardware_row_simple(node_name, divre, component_type, slot, part_number, serial_number,
                              model_description, version, operational_status, remarks, wb_obj):
    ws = wb_obj[HARDWARE_SHEET]
    row = ws.max_row + 1
    thin = THIN_BORDER
    entries = [
        (f'A{row}', str(row-5), 'center', False),
        (f'B{row}', node_name, 'center', True),
        (f'C{row}', divre, 'center', False),
        (f'D{row}', component_type or '', 'center', True),
        (f'E{row}', slot or '', 'center', True),
        (f'F{row}', part_number or '', 'center', False),
        (f'G{row}', serial_number or '', 'center', False),
        (f'H{row}', model_description or '', 'left', True),
        (f'I{row}', version or '', 'center', False),
        (f'J{row}', operational_status or 'Online', 'center', False),
        (f'K{row}', remarks or '', 'left', True),
    ]
    for addr, val, hc, wrap in entries:
        ws[addr] = val
        ws[addr].alignment = Alignment(horizontal=hc, wrap_text=wrap)
        ws[addr].border = thin
        ws[addr].font = Font(name='Calibri')
    ws[f'B{row}'].font = Font(name='Calibri', bold=True, color='2E4A6B')
    logger.debug(f"write_hardware_row_simple node={node_name} component={component_type} slot={slot}")


def write_data_row_simple(node_name, divre, desc_interface, iface_name, module_type,
                           port_capacity, current_traffic_gb, current_utilization, traffic_alert, wb_obj):
    ws = wb_obj[MAIN_SHEET]; row = ws.max_row + 1; thin = THIN_BORDER
    ws[f'A{row}'] = str(row - 5); ws[f'A{row}'].alignment = Alignment(horizontal='center'); ws[f'A{row}'].border = thin; ws[f'A{row}'].font=Font(name='Calibri')
    ws[f'B{row}'] = node_name; ws[f'B{row}'].alignment = Alignment(horizontal='center', wrap_text=True); ws[f'B{row}'].font = Font(name='Calibri', bold=True, color='2E4A6B'); ws[f'B{row}'].border = thin
    ws[f'C{row}'] = divre; ws[f'C{row}'].alignment = Alignment(horizontal='center'); ws[f'C{row}'].font=Font(name='Calibri'); ws[f'C{row}'].border = thin
    ws[f'D{row}'] = desc_interface; ws[f'D{row}'].alignment = Alignment(horizontal='left', wrap_text=True); ws[f'D{row}'].font=Font(name='Calibri'); ws[f'D{row}'].border = thin
    ws[f'E{row}'] = iface_name; ws[f'E{row}'].alignment = Alignment(horizontal='center'); ws[f'E{row}'].font=Font(name='Calibri'); ws[f'E{row}'].border = thin
    ws[f'F{row}'] = module_type; ws[f'F{row}'].alignment = Alignment(horizontal='center', wrap_text=True); ws[f'F{row}'].font=Font(name='Calibri'); ws[f'F{row}'].border = thin
    ws[f'G{row}'] = port_capacity; ws[f'G{row}'].alignment = Alignment(horizontal='center'); ws[f'G{row}'].font=Font(name='Calibri'); ws[f'G{row}'].border = thin
    try:
        v = float(current_traffic_gb)
        disp = f"{v:.2f} GB" if v >= 1.0 else f"{v*1024:.2f} MB" if v*1024 >= 1.0 else f"{int(round(v*(1024**3)))} B"
    except Exception:
        disp = "0.00 GB"
    ws[f'H{row}'] = disp; ws[f'H{row}'].alignment = Alignment(horizontal='center'); ws[f'H{row}'].font=Font(name='Calibri'); ws[f'H{row}'].border = thin
    try:
        util_fraction = float(current_utilization) if float(current_utilization) <= 1.0 else float(current_utilization)/100.0
    except Exception:
        util_fraction = 0.0
    ws[f'I{row}'] = util_fraction; ws[f'I{row}'].number_format = '0.00%'; ws[f'I{row}'].alignment = Alignment(horizontal='center'); ws[f'I{row}'].font = Font(name='Calibri', color=GREEN); ws[f'I{row}'].border = thin
    ws[f'J{row}'] = _status_color(util_fraction); ws[f'J{row}'].alignment = Alignment(horizontal='center'); ws[f'J{row}'].font=Font(name='Calibri'); ws[f'J{row}'].border = thin
    logger.debug(f"write_data_row_simple node={node_name} iface={iface_name} util={util_fraction}")


def write_utilisasi_port_row_simple(node_name, divre, iface_name, module_type, port_capacity,
                                    last_flapped, sfp_present, configured, desc_interface, status, flap_alert, wb_obj):
    ws = wb_obj[UTIL_SHEET]; row = ws.max_row + 1; thin = THIN_BORDER
    def _w(addr, val, hc='center', wrap=False):
        ws[addr]=val; ws[addr].alignment=Alignment(horizontal=hc, wrap_text=wrap); ws[addr].border=thin; ws[addr].font=Font(name='Calibri')
    _w(f'A{row}', str(row-5)); _w(f'B{row}', node_name, 'center', True); ws[f'B{row}'].font = Font(name='Calibri', bold=True, color='2E4A6B')
    _w(f'C{row}', divre); _w(f'D{row}', iface_name)
    _w(f'E{row}', module_type, 'center', True); _w(f'F{row}', port_capacity)
    _w(f'G{row}', last_flapped or 'N/A', 'center', True); _w(f'H{row}', sfp_present or 'Unknown')
    _w(f'I{row}', configured or 'No'); _w(f'J{row}', desc_interface or '', 'left', True)
    _w(f'K{row}', status or 'UNUSED'); _w(f'L{row}', flap_alert or 'Stable', 'center', True)
    alert_updown = 'Down'
    try:
        main_ws = wb_obj[MAIN_SHEET]
        exists = {(str(main_ws[f'B{r}'].value or '').strip().lower(), str(main_ws[f'E{r}'].value or '').strip().lower()) for r in range(6, main_ws.max_row+1)}
        if (str(node_name).strip().lower(), str(iface_name).strip().lower()) in exists: alert_updown = 'Up'
    except Exception:
        alert_updown = 'Down'
    _w(f'M{row}', alert_updown)
    logger.debug(f"write_utilisasi_port_row_simple node={node_name} iface={iface_name} status={status} updown={alert_updown}")


def write_alarm_row_simple(node_name, divre, alarm_time, alarm_type, alarm_desc, severity, status, wb_obj):
    ws = wb_obj[ALARM_SHEET]; row = ws.max_row + 1; thin = THIN_BORDER
    for addr, val, hc, wrap in ((f'A{row}', str(row-5), 'center', False), (f'B{row}', node_name, 'center', True), (f'C{row}', divre, 'center', False), (f'D{row}', alarm_time or 'N/A', 'center', False), (f'E{row}', alarm_type or 'Status', 'center', False), (f'F{row}', alarm_desc or 'No alarms currently active', 'left', True), (f'G{row}', severity or 'System', 'center', False), (f'H{row}', status or 'No Active', 'center', False)):
        ws[addr]=val; ws[addr].alignment=Alignment(horizontal=hc, wrap_text=wrap); ws[addr].border=THIN_BORDER; ws[addr].font=Font(name='Calibri')
    ws[f'B{row}'].font = Font(name='Calibri', bold=True, color='2E4A6B')
    logger.debug(f"write_alarm_row_simple node={node_name} alarm_type={alarm_type} severity={severity}")

# ---------- XML helpers & parsers ----------
def _extract_xml_fragment(buff):
    try:
        if not buff: return ''
        b = str(buff)
        s = b.find('<rpc-reply'); e = b.rfind('</rpc-reply>')
        frag = b[s:e+len('</rpc-reply>')] if s!=-1 and e!=-1 else b
        repaired = _repair_corrupt_xml(frag)
        logger.debug(f"_extract_xml_fragment len_in={len(b)} len_out={len(repaired)}")
        return repaired
    except Exception:
        return str(buff) if buff else ''


def _parse_fragments_to_dom(fragment, tag_hint=None):
    if not fragment: return None
    try:
        doc = minidom.parseString(fragment); return doc
    except Exception:
        try:
            return minidom.parseString(f"<root>{fragment}</root>")
        except Exception:
            return None


def sanitize_xml_text(raw: str) -> str:
    if not raw: return ''
    s = str(raw)
    # Remove CLI echoes if present
    s = re.sub(r'(?:^\s*set\s+cli\s+screen-length.*\nshow\s+.*\nfile\s+show\s+/var/tmp/.*)\s*$', '', s, flags=re.IGNORECASE|re.MULTILINE)
    s0 = s.find('<rpc-reply'); e0 = s.rfind('</rpc-reply>')
    if s0!=-1 and e0!=-1:
        repaired = _repair_corrupt_xml(s[s0:e0+len('</rpc-reply>')])
        logger.debug(f"sanitize_xml_text rpc-reply repaired len={len(repaired)}")
        return repaired
    m = re.search(r'<interface-information[\s\S]*?</interface-information>', s, flags=re.IGNORECASE)
    repaired = _repair_corrupt_xml(m.group(0)) if m else _repair_corrupt_xml(s)
    logger.debug(f"sanitize_xml_text generic repaired len={len(repaired)}")
    return repaired


def _repair_corrupt_xml(x: str) -> str:
    if not x: return ''
    # strip ANSI
    x = re.sub(r"\x1B\[[0-?]*[ -/]*[@-~]", "", x)
    # strip control chars
    x = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", " ", x)
    if '<rpc-reply' in x and '</rpc-reply>' not in x:
        x += '</rpc-reply>'
    try:
        minidom.parseString(x)
        return x
    except Exception:
        try:
            minidom.parseString(f"<root>{x}</root>")
            return f"<root>{x}</root>"
        except Exception:
            s = x.find('<'); e = x.rfind('>')
            if s!=-1 and e!=-1:
                y = x[s:e+1]
                try:
                    minidom.parseString(y)
                    return y
                except Exception:
                    return y
            return x


def _get_first_text(parent, tag):
    try:
        nodes = parent.getElementsByTagName(tag)
        return nodes[0].firstChild.data.strip() if nodes and nodes[0].firstChild else ''
    except Exception: return ''

# ---------- Interface parsers ----------
def parse_interfaces_xml_basic(xml_text: str):
    results = []
    if not xml_text: return results
    try:
        frag = _extract_xml_fragment(xml_text); doc = _parse_fragments_to_dom(frag)
        phys_list = doc.getElementsByTagName('physical-interface') if doc else []
        if not phys_list: return results
        for phys in phys_list:
            name = desc = speed = last_flapped = ''
            in_bps = out_bps = 0
            n_nodes = phys.getElementsByTagName('name'); d_nodes = phys.getElementsByTagName('description'); s_nodes = phys.getElementsByTagName('speed')
            if n_nodes and n_nodes[0].firstChild: name = n_nodes[0].firstChild.data.strip()
            if d_nodes and d_nodes[0].firstChild: desc = d_nodes[0].firstChild.data.strip()
            if s_nodes and s_nodes[0].firstChild: speed = s_nodes[0].firstChild.data.strip()
            ts_nodes = phys.getElementsByTagName('traffic-statistics')
            if ts_nodes:
                in_nodes = ts_nodes[0].getElementsByTagName('input-bps'); out_nodes= ts_nodes[0].getElementsByTagName('output-bps')
                try: in_bps = int(in_nodes[0].firstChild.data.strip()) if in_nodes and in_nodes[0].firstChild else 0
                except: in_bps = 0
                try: out_bps= int(out_nodes[0].firstChild.data.strip()) if out_nodes and out_nodes[0].firstChild else 0
                except: out_bps = 0
            lf_nodes = phys.getElementsByTagName('interface-flapped')
            if lf_nodes and lf_nodes[0].firstChild: last_flapped = lf_nodes[0].firstChild.data.strip()
            s_up = (speed or '').upper()
            cap = '100Gbps' if '100G' in s_up else '10Gbps' if '10G' in s_up else '1Gbps' if ('1G' in s_up or '1000M' in s_up) else (speed or '')
            cap_bps = 100_000_000_000 if cap=='100Gbps' else 10_000_000_000 if cap=='10Gbps' else 1_000_000_000 if cap=='1Gbps' else 0
            peak = max(in_bps, out_bps); util = (peak / cap_bps) if (cap_bps > 0 and peak > 0) else 0.0; traffic_gb = peak / (1024**3)
            if name and (name.startswith('lc-') or name.startswith('pfe-') or name.startswith('pfh-')): continue
            if name and name.startswith(('ae-','et-','xe-','ge-')):
                results.append({'iface': name, 'desc': desc, 'capacity': cap or '', 'traffic_gb': traffic_gb, 'util': util, 'last_flapped': last_flapped})
        logger.debug(f"parse_interfaces_xml_basic parsed={len(results)}")
        return results
    except Exception as e:
        append_error_log(get_debug_log_path('interfaces_parse_errors.log'), f'Basic interfaces parse failed: {e}')
        return results

# ---------- Alarms & hardware parsers ----------
def parse_chassis_alarms(xml_text: str):
    alarms = []
    if not xml_text: return alarms
    try:
        frag = sanitize_xml_text(xml_text); doc = _parse_fragments_to_dom(frag)
        if not doc: return alarms
        for det in doc.getElementsByTagName('alarm-detail'):
            t=_get_first_text(det,'alarm-time'); typ=_get_first_text(det,'alarm-type')
            desc=_get_first_text(det,'alarm-short-description') or _get_first_text(det,'alarm-description')
            sev=_get_first_text(det,'alarm-class'); stat='Active'
            if any([t,typ,desc,sev]): alarms.append({'time': t, 'type': typ, 'description': desc, 'severity': sev, 'status': stat})
        logger.debug(f"parse_chassis_alarms parsed={len(alarms)}")
        return alarms
    except Exception as e:
        append_error_log(get_debug_log_path('alarms_parse_errors.log'), f'Parse alarms failed: {e}')
        return alarms


def parse_hardware_inventory(xml_text: str):
    items = []
    if not xml_text: return items
    try:
        frag = sanitize_xml_text(xml_text); doc = _parse_fragments_to_dom(frag)
        if not doc: return items
        for mod in doc.getElementsByTagName('chassis-module'):
            name = _get_first_text(mod, 'name') or 'Module'
            part = _get_first_text(mod, 'part-number'); serial= _get_first_text(mod, 'serial-number')
            desc = _get_first_text(mod, 'description') or _get_first_text(mod, 'model-number') or ''
            ver = _get_first_text(mod, 'version'); clei = _get_first_text(mod, 'clei-code')
            state = _get_first_text(mod, 'state'); temp = _get_first_text(mod, 'temperature')
            comments = []
            if _get_first_text(mod,'model-number'): comments.append(f"Model: {_get_first_text(mod,'model-number')}")
            if clei: comments.append(f"CLEI: {clei}")
            if state: comments.append(f"State: {state}")
            if temp: comments.append(f"Temp: {temp}°C")
            items.append({'component_type': name.split()[0] if name else 'Module',
                         'slot': name, 'part': part, 'serial': serial, 'model': desc,
                         'version': ver, 'status': state or 'Online', 'comments': ", ".join([c for c in comments if c])})
        for fpc in doc.getElementsByTagName('fpc'):
            slot = _get_first_text(fpc, 'slot'); state= _get_first_text(fpc, 'state') or 'Online'
            temp = _get_first_text(fpc, 'temperature'); part = _get_first_text(fpc, 'part-number')
            serial = _get_first_text(fpc, 'serial-number')
            model = _get_first_text(fpc, 'description') or _get_first_text(fpc, 'model-number')
            ver = _get_first_text(fpc, 'version')
            comments = []
            if temp: comments.append(f"Temp: {temp}°C")
            if state: comments.append(f"State: {state}")
            items.append({'component_type':'FPC','slot': f"FPC {slot}" if slot else 'FPC','part':part,'serial':serial,
                         'model':model,'version':ver,'status':state,'comments':", ".join(comments)})
        logger.debug(f"parse_hardware_inventory parsed={len(items)}")
        return items
    except Exception as e:
        append_error_log(get_debug_log_path('hardware_parse_errors.log'), f'Parse hardware failed: {e}')
        return items

# ---------- Build map ----------
def build_fpc_model_map(hw_items):
    fpc_map = {}
    for it in (hw_items or []):
        if str(it.get('component_type','')).upper() == 'FPC':
            slot_str = it.get('slot','')
            m = re.search(r'(?i)FPC\s*(\d+)', slot_str)
            if m:
                slot = int(m.group(1))
                model = it.get('model') or 'FPC'
                fpc_map[slot] = model
    logger.debug(f"build_fpc_model_map entries={len(fpc_map)}")
    return fpc_map

# ---------- System Performance writer ----------
def worksheet_system_performance(ws, system_data=None):
    if system_data is None: system_data = {}
    row = 6; counter = 1
    for node, info in system_data.items():
        if not isinstance(info, dict): continue
        platform = info.get('platform', 'mx960'); current_sw = info.get('current_sw','')
        current_sw_type = info.get('current_sw_type','JUNOS')
        current_sw_disp = (current_sw or current_sw_type)
        loopback = info.get('loopback_address','-'); mem_util = info.get('memory_util',3)
        cpu_usage = info.get('cpu_usage',5); total_space= info.get('total_space',54272)
        used_space = info.get('used_space',6144); free_space = info.get('free_space',48128)
        disk_util = info.get('disk_util', int(round((used_space/total_space)*100)) if total_space else 0)
        temperature= info.get('temperature',49)
        mem_rec = info.get('memory_recommendation','NORMAL - Optimal Performance')
        cpu_rec = info.get('cpu_recommendation','NORMAL - Optimal Performance')
        disk_rec = info.get('disk_recommendation','NORMAL - Adequate Free Space')
        data_row=[counter,get_area_pop_from_node(node),node,loopback,'ACTIVE',current_sw_disp,platform,
                  f"{mem_util}%",mem_rec,f"{cpu_usage}%",cpu_rec,
                  total_space,used_space,free_space,f"{disk_util}%",disk_rec,f"{temperature}°C"]
        for cidx,val in enumerate(data_row,1):
            cell = ws.cell(row=row,column=cidx); cell.value=val
            cell.font = Font(name='Calibri')
            if cidx in (12,13,14): cell.number_format = '#,##0'
            if cidx in (8,10,15): cell.font = Font(name='Calibri', color=GREEN)
            wrap_cols = {3,9,11,16}
            cell.alignment=Alignment(horizontal='center',vertical='center', wrap_text=(cidx in wrap_cols))
            cell.border = THIN_BORDER
        row+=1; counter+=1
    style_data_rows(ws, start_row=6, wrap_cols={'C','I','K','P'})
    logger.debug(f"worksheet_system_performance rows_written={counter-1}")

# ---------- Dashboard synthesis ----------
def _collect_top_interfaces(wb, top_n=TOP_N):
    ws = wb[MAIN_SHEET]; data = []
    for r in range(6, ws.max_row+1):
        node = ws[f'B{r}'].value; iface = ws[f'E{r}'].value; util = ws[f'I{r}'].value or 0.0
        module = ws[f'F{r}'].value; bw = ws[f'G{r}'].value; status = ws[f'J{r}'].value
        try: util_frac = float(util)
        except: util_frac = 0.0
        data.append((util_frac, node, iface, module, bw, status))
    data.sort(reverse=True, key=lambda x: x[0]); logger.debug(f"_collect_top_interfaces count={len(data)} top_n={top_n}"); return data[:top_n]


def _count_flap_alerts(wb):
    ws = wb[UTIL_SHEET]
    counts = {'CRITICAL':0,'WARNING':0,'INFO':0,'NORMAL':0}
    for r in range(6, ws.max_row+1):
        alert = str(ws[f'L{r}'].value or '').upper()
        if 'RECENT FLAP - <=' in alert and '5MIN' in alert: counts['CRITICAL']+=1
        elif 'RECENT FLAP - <=' in alert and '30MIN' in alert: counts['WARNING']+=1
        elif 'FLAPPED - <=' in alert and '2H' in alert: counts['INFO']+=1
        else: counts['NORMAL']+=1
    logger.debug(f"_count_flap_alerts counts={counts}")
    return counts


def populate_dashboard_like_example(wb, nodes):
    ws = wb[DASHBOARD_SHEET]
    total_nodes = len(nodes)
    try:
        ws_main = wb[MAIN_SHEET]; active_ifaces = _count_data_rows(ws_main, start_row=6, must_have_cols=(2,))
    except Exception: active_ifaces = 0
    try:
        ws_hw = wb[HARDWARE_SHEET]; hw_items = _count_data_rows(ws_hw, start_row=6, must_have_cols=(2,))
    except Exception: hw_items = 0
    try:
        ws_alarm = wb[ALARM_SHEET]; alarms_rows = _count_data_rows(ws_alarm, start_row=6, must_have_cols=(2,))
    except Exception: alarms_rows = 0

    def setv(c,v,color=None): ws[c]=v; ws[c].alignment = Alignment(horizontal='center', wrap_text=True); ws[c].border = THIN_BORDER; ws[c].font = Font(name='Calibri', bold=True, color=(color or None))
    setv('C6', total_nodes); setv('D6', 'Normal', GREEN)
    setv('C7', active_ifaces); setv('D7', 'Active', GREEN)
    setv('C8', hw_items); setv('D8', 'Online', GREEN)
    setv('C9', alarms_rows); setv('D9', 'No Alarms' if alarms_rows==0 else f'{alarms_rows} Active', GREEN)

    start = ws.max_row + 2
    ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=8)
    ws.cell(row=start, column=2).value = 'TOP INTERFACE UTILIZATION'
    ws.cell(row=start, column=2).font = Font(name='Calibri', bold=True, color=WHITE); ws.cell(row=start, column=2).alignment = Alignment(horizontal='center', wrap_text=True)
    ws.cell(row=start, column=2).fill = fill(ORANGE); ws.row_dimensions[start].height=24
    headers = ['Node Name','Interface','Utilization %','Module Type','Bandwidth','Status']
    for i,h in enumerate(headers,2):
        cell = ws.cell(row=start+1, column=i, value=headers[i-2])
        cell.font = Font(name='Calibri', bold=True, color=WHITE); cell.fill = fill(ORANGE)
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = THIN_BORDER
    top = _collect_top_interfaces(wb, TOP_N)
    rr = start+2
    for util_frac, node, iface, module, bw, status in top:
        ws.cell(row=rr, column=2, value=node or '').border = THIN_BORDER
        ws.cell(row=rr, column=3, value=iface or '').border = THIN_BORDER
        c_util = ws.cell(row=rr, column=4, value=util_frac); c_util.number_format = '0.00%'; c_util.font = Font(name='Calibri', color=GREEN)
        ws.cell(row=rr, column=5, value=module or '').border = THIN_BORDER
        ws.cell(row=rr, column=6, value=bw or '').border = THIN_BORDER
        ws.cell(row=rr, column=7, value='Active' if status=='Green' else 'Watch' if status=='Yellow' else 'Hot').border = THIN_BORDER
        band = fill(ZEBRA_A if (rr-(start+2))%2==0 else ZEBRA_B)
        for cc in range(2,8): ws.cell(row=rr, column=cc).fill = band
        rr += 1

    start2 = rr + 2
    ws.merge_cells(start_row=start2, start_column=2, end_row=start2, end_column=8)
    ws.cell(row=start2, column=2).value = 'INTERFACE FLAP ALERT SUMMARY'
    ws.cell(row=start2, column=2).font = Font(name='Calibri', bold=True, color=WHITE); ws.cell(row=start2, column=2).alignment = Alignment(horizontal='center', wrap_text=True)
    ws.cell(row=start2, column=2).fill = fill(RED); ws.row_dimensions[start2].height=24
    headers2 = ['Alert Level','Count','Status','Last Critical','Stability','Action']
    for i,h in enumerate(headers2,2):
        cell = ws.cell(row=start2+1, column=i, value=headers2[i-2])
        cell.font = Font(name='Calibri', bold=True, color=WHITE); cell.fill = fill(RED)
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = THIN_BORDER
    c = _count_flap_alerts(wb)
    mapping = [('CRITICAL', 'None', 'Immediate','Investigate'), ('WARNING', 'None', 'Monitor', 'Review'), ('INFO', 'None', 'Normal', 'Track'), ('NORMAL', 'Stable','Excellent','Continue')]
    rr2 = start2+2
    for lvl, last, stability, action in mapping:
        ws.cell(row=rr2, column=2, value=lvl).border = THIN_BORDER
        ws.cell(row=rr2, column=3, value=c.get(lvl,0)).border = THIN_BORDER
        ws.cell(row=rr2, column=4, value='None' if c.get(lvl,0)==0 else 'Active').border = THIN_BORDER
        ws.cell(row=rr2, column=5, value=last).border = THIN_BORDER
        ws.cell(row=rr2, column=6, value=stability).border = THIN_BORDER
        ws.cell(row=rr2, column=7, value=action).border = THIN_BORDER
        band = fill(ZEBRA_A if (rr2-(start2+2))%2==0 else ZEBRA_B)
        for cc in range(2,8): ws.cell(row=rr2, column=cc).fill = band
        rr2 += 1

    start3 = rr2 + 2
    ws.merge_cells(start_row=start3, start_column=2, end_row=start3, end_column=8)
    ws.cell(row=start3, column=2).value = 'SYSTEM STATUS SUMMARY'
    ws.cell(row=start3, column=2).font = Font(name='Calibri', bold=True, color=WHITE); ws.cell(row=start3, column=2).alignment = Alignment(horizontal='center', wrap_text=True)
    ws.cell(row=start3, column=2).fill = fill(GREEN); ws.row_dimensions[start3].height=24
    headers3 = ['Component','Total','Online','Health Status','Status','Action']
    for i,h in enumerate(headers3,2):
        cell = ws.cell(row=start3+1, column=i, value=headers3[i-2])
        cell.font = Font(name='Calibri', bold=True, color=WHITE); cell.fill = fill(GREEN)
        cell.alignment = Alignment(horizontal='center', wrap_text=True); cell.border = THIN_BORDER
    rr3 = start3+2
    rows = [
        ('Network', len(nodes), len(nodes), '✓ 100%','Online','Monitor'),
        ('Interfaces', _count_data_rows(wb[MAIN_SHEET], start_row=6, must_have_cols=(2,)), _count_data_rows(wb[MAIN_SHEET], start_row=6, must_have_cols=(2,)), '✓ 100%','Active','Normal'),
        ('Hardware', _count_data_rows(wb[HARDWARE_SHEET], start_row=6, must_have_cols=(2,)), _count_data_rows(wb[HARDWARE_SHEET], start_row=6, must_have_cols=(2,)), '✓ 100%','Operational','Good'),
        ('Alarms', _count_data_rows(wb[ALARM_SHEET], start_row=6, must_have_cols=(2,)), _count_data_rows(wb[ALARM_SHEET], start_row=6, must_have_cols=(2,)), 'Check' if _count_data_rows(wb[ALARM_SHEET], start_row=6, must_have_cols=(2,))>0 else 'Clear','Alert' if _count_data_rows(wb[ALARM_SHEET], start_row=6, must_have_cols=(2,))>0 else 'Clear','Monitor')
    ]
    for comp, tot, onl, health, status, action in rows:
        ws.cell(row=rr3, column=2, value=comp).border = THIN_BORDER
        ws.cell(row=rr3, column=3, value=tot).border = THIN_BORDER
        ws.cell(row=rr3, column=4, value=onl).border = THIN_BORDER
        hc = ws.cell(row=rr3, column=5, value=health); hc.font = Font(name='Calibri', color=GREEN); hc.border = THIN_BORDER
        ws.cell(row=rr3, column=6, value=status).border = THIN_BORDER
        ws.cell(row=rr3, column=7, value=action).border = THIN_BORDER
        band = fill(ZEBRA_A if (rr3-(start3+2))%2==0 else ZEBRA_B)
        for cc in range(2,8): ws.cell(row=rr3, column=cc).fill = band
        rr3 += 1

    start4 = rr3 + 2
    ws.merge_cells(start_row=start4, start_column=2, end_row=start4, end_column=8)
    ws.cell(row=start4, column=2).value = 'RECOMMENDATIONS & INSIGHTS'
    ws.cell(row=start4, column=2).font = Font(name='Calibri', bold=True, color=WHITE); ws.cell(row=start4, column=2).alignment = Alignment(horizontal='center', wrap_text=True)
    ws.cell(row=start4, column=2).fill = fill(PURPLE)
    ws.merge_cells(start_row=start4+1, start_column=2, end_row=start4+1, end_column=8)
    ws.cell(row=start4+1, column=2).value = 'Recommendations will be populated automatically based on data analysis.'
    ws.cell(row=start4+1, column=2).alignment = Alignment(horizontal='left', wrap_text=True)
    logger.debug("populate_dashboard_like_example completed")

# ---------- finalize & conditional ----------
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
RED_FILL = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
BLUE_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
GREY_FILL = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

def finalize_tables(wb):
    try: style_data_rows(wb[MAIN_SHEET], start_row=6, wrap_cols={'B','D','F'})
    except Exception: pass
    try: style_data_rows(wb[UTIL_SHEET], start_row=6, wrap_cols={'B','E','G','J','L'})
    except Exception: pass
    try: style_data_rows(wb[ALARM_SHEET], start_row=6, wrap_cols={'B','F'})
    except Exception: pass
    try: style_data_rows(wb[HARDWARE_SHEET], start_row=6, wrap_cols={'B','D','E','H','K'})
    except Exception: pass
    for sheet, rng_end_col in ((MAIN_SHEET,'J'), (UTIL_SHEET,'M'), (ALARM_SHEET,'H'), (HARDWARE_SHEET,'K'), (SYSTEM_SHEET,'Q')):
        try: ws = wb[sheet]
        except Exception: ws = None
        if ws and ws.max_row>5:
            ws.auto_filter.ref = f"A5:{rng_end_col}{ws.max_row}"
            dynamic_resize_columns(ws, header_row=5, start_row=6)
    logger.debug("finalize_tables completed")


def apply_conditional_formatting(wb):
    try:
        ws = wb[MAIN_SHEET]
        max_row = ws.max_row
        if max_row > 6:
            rngI = f"I6:I{max_row}"
            ws.conditional_formatting.add(rngI, CellIsRule(operator='lessThan', formula=['0.60'], stopIfTrue=False, fill=GREEN_FILL))
            ws.conditional_formatting.add(rngI, CellIsRule(operator='between', formula=['0.60','0.80'], stopIfTrue=False, fill=YELLOW_FILL))
            ws.conditional_formatting.add(rngI, CellIsRule(operator='greaterThanOrEqual', formula=['0.80'], stopIfTrue=False, fill=RED_FILL))
            rngJ = f"J6:J{max_row}"
            ws.conditional_formatting.add(rngJ, FormulaRule(formula=[f"$I6<0.60"], stopIfTrue=False, fill=GREEN_FILL))
            ws.conditional_formatting.add(rngJ, FormulaRule(formula=[f"AND($I6>=0.60,$I6<=0.80)"], stopIfTrue=False, fill=YELLOW_FILL))
            ws.conditional_formatting.add(rngJ, FormulaRule(formula=[f"$I6>=0.80"], stopIfTrue=False, fill=RED_FILL))
    except Exception:
        pass
    try:
        ws2 = wb[UTIL_SHEET]; max_row2 = ws2.max_row
        if max_row2 > 6:
            rng_flap = f"L6:L{max_row2}"
            ws2.conditional_formatting.add(rng_flap, FormulaRule(formula=['NOT(ISERROR(SEARCH("RECENT FLAP - <=5min",L6)))'], stopIfTrue=False, fill=RED_FILL))
            ws2.conditional_formatting.add(rng_flap, FormulaRule(formula=['NOT(ISERROR(SEARCH("Recent flap - <=30min",L6)))'], stopIfTrue=False, fill=YELLOW_FILL))
            ws2.conditional_formatting.add(rng_flap, FormulaRule(formula=['NOT(ISERROR(SEARCH("Flapped - <=2h",L6)))'], stopIfTrue=False, fill=BLUE_FILL))
            ws2.conditional_formatting.add(rng_flap, FormulaRule(formula=['NOT(ISERROR(SEARCH("Stable",L6)))'], stopIfTrue=False, fill=GREEN_FILL))
            rng_updown = f"M6:M{max_row2}"
            ws2.conditional_formatting.add(rng_updown, FormulaRule(formula=['NOT(ISERROR(SEARCH("Up",M6)))'], stopIfTrue=False, fill=GREEN_FILL))
            ws2.conditional_formatting.add(rng_updown, FormulaRule(formula=['NOT(ISERROR(SEARCH("Down",M6)))'], stopIfTrue=False, fill=GREY_FILL))
    except Exception:
        pass
    logger.debug("apply_conditional_formatting completed")

# ---------- Junos CLI session ----------

# --- Platform & SW detection helper (EVO-aware + RAW version token) ---
import re

PLATFORM_TAGS = {
    'PTX10004': ['PTX10004','JNP10004','JNP10K','JNP10K-LC1202','JNP10K-RE1','JNP10004-CHAS'],
    'MX960': ['MX960','CHAS-MX960','ENHANCED MX960 MIDPLANE'],
    'MX480': ['MX480','CHAS-MX480','ENHANCED MX480 MIDPLANE'],
    'MX240': ['MX240','CHAS-MX240'],
    'MX204': ['MX204'],
    'MX304': ['MX304','CHAS-MX304'],
    'MX2008': ['MX2008'],
    'MX10003': ['MX10003'],
}

def _extract_raw_junos_version(show_version_text: str) -> str:
    """Return ONLY version token: e.g., '19.4R3-S7.3' or '21.4R3-S1.6-EVO'."""
    if not show_version_text:
        return ''
    s = str(show_version_text)
    # 1) Prefer 'Junos:' line
    for line in s.splitlines():
        if re.match(r'^\s*Junos\s*:\s*', line, flags=re.IGNORECASE):
            token = re.sub(r'^\s*Junos\s*:\s*', '', line, flags=re.IGNORECASE).strip()
            m = re.search(r'\b(\d+\.\d+[A-Za-z0-9\.\-]*R\d+(?:[A-Za-z0-9\.\-]*)?)\b', token)
            if m:
                return m.group(1)
    # 2) Fallback: look anywhere (covers '...-21.4R3-S1.6-EVO')
    m2 = re.search(r'\b(\d+\.\d+[A-Za-z0-9\.\-]*R\d+(?:[A-Za-z0-9\.\-]*)?)\b', s, flags=re.IGNORECASE)
    return m2.group(1) if m2 else ''

def _detect_platform_and_sw(node: str, hw_items: list, show_version_text: str):
    """Return (platform, os_type, raw_version_token)."""
    def U(x): return (str(x or '')).upper()
    platform=None
    # 1) hardware tags
    try:
        for it in hw_items or []:
            comp=U(it.get('component_type')); model=U(it.get('model')); slot=U(it.get('slot'))
            blob = ' '.join([model,comp,slot])
            for plat,tags in PLATFORM_TAGS.items():
                if any(tag in blob for tag in tags): platform=plat; break
            if platform: break
            if ('MIDPLANE' in comp) or ('CHASSIS' in comp):
                for plat,tags in PLATFORM_TAGS.items():
                    if any(tag in blob for tag in tags): platform=plat; break
            if platform: break
    except Exception:
        pass
    # 2) 'Model:' from show version
    if not platform:
        sv=U(show_version_text)
        m=re.search(r'\bMODEL\s*:\s*([A-Z0-9\-]+)\b', sv)
        if m:
            model_tag=m.group(1)
            for plat,tags in PLATFORM_TAGS.items():
                if any(model_tag.startswith(tag) for tag in tags): platform=plat; break
            if not platform:
                if re.match(r'^PTX\d+$', model_tag): platform=model_tag
                elif re.match(r'^MX\d+$', model_tag): platform=model_tag
    # 3) hostname fallback
    if not platform:
        hn=U(node)
        m=re.search(r'(PTX\d+)', hn)
        if m: platform=m.group(1)
        else:
            m2=re.search(r'(MX\d+)', hn)
            platform = m2.group(1) if m2 else None
    if not platform: platform='Unknown'

    # OS type: EVO vs JUNOS (robust tokens)
    os_type='JUNOS'
    try:
        sup=U(show_version_text)
        if ('EVOLVED' in sup) or ('JUNOS-EVO' in sup) or ('-EVO' in sup) or ('JUNOS EVO' in sup) or ('JUNOS OS EVO' in sup):
            os_type='Junos EVO'
        elif 'JUNOS' in sup:
            os_type='JUNOS'
    except Exception:
        pass
    raw=_extract_raw_junos_version(show_version_text)
    return platform, os_type, raw



# --- System Performance parsers (CPU/Memory/Storage)
import json

def _to_mb(val_str: str) -> float:
    s = str(val_str or '').strip()
    if not s:
        return 0.0
    try:
        if re.match(r'^\d+(?:\.\d+)?$', s):
            return float(s)
        m = re.match(r'^(\d+(?:\.\d+)?)([KkMmGgTt])', s)
        if m:
            num = float(m.group(1)); unit = m.group(2).upper()
            factor = {'K': 1/1024.0, 'M': 1.0, 'G': 1024.0, 'T': 1024.0*1024.0}[unit]
            return num * factor
    except Exception:
        pass
    return 0.0

def _parse_cpu_used_from_re(show_re_text: str):
    if not show_re_text:
        return None
    idle_candidates = []
    for line in str(show_re_text).splitlines():
        if 'CPU utilization:' in line or re.search(r'\bIdle\b', line, flags=re.I):
            m = re.search(r'Idle\s+(\d+)', line)
            if m:
                try:
                    idle_candidates.append(int(m.group(1)))
                except Exception:
                    pass
    if idle_candidates:
        idle = max(idle_candidates)
        return max(0, min(100, 100 - idle))
    return None

def _parse_mem_util_from_re_or_sysmem(show_re_text: str, show_mem_text: str):
    if show_re_text:
        for line in str(show_re_text).splitlines():
            m = re.search(r'Memory\s+utilization\s+(\d+)\s*percent', line, flags=re.I)
            if m:
                try:
                    return float(m.group(1))
                except Exception:
                    pass
    total = reserved = free = cache = inactive = None
    for line in str(show_mem_text or '').splitlines():
        m_tot = re.search(r'\bTotal\s+memory\s*:\s*(\d+)\s*Kbytes', line, flags=re.I)
        if m_tot:
            total = float(m_tot.group(1)) / 1024.0; continue
        m_res = re.search(r'\bReserved\s+memory\s*:\s*(\d+)\s*Kbytes', line, flags=re.I)
        if m_res:
            reserved = float(m_res.group(1)) / 1024.0; continue
        m_free = re.search(r'\bFree\s+memory\s*:\s*(\d+)\s*Kbytes', line, flags=re.I)
        if m_free:
            free = float(m_free.group(1)) / 1024.0; continue
        m_cache = re.search(r'\bCache\s+memory\s*:\s*(\d+)\s*Kbytes', line, flags=re.I)
        if m_cache:
            cache = float(m_cache.group(1)) / 1024.0; continue
        m_inact = re.search(r'\bInactive\s+memory\s*:\s*(\d+)\s*Kbytes', line, flags=re.I)
        if m_inact:
            inactive = float(m_inact.group(1)) / 1024.0; continue
    if total:
        total_re = (total or 0) + (reserved or 0)
        used_calc = (total_re or 0) - (free or 0) - (cache or 0) - (inactive or 0)
        try:
            mu = (used_calc * 100.0) / max(total_re, 1.0)
            return float(round(mu))
        except Exception:
            return None
    return None

PREF_MOUNT_PATTERNS = [r'/(?:\\.mount/)?var\\b', r'/var\\b']

def _parse_storage_xml(xml_text: str):
    try:
        if not xml_text:
            return None
        frag = sanitize_xml_text(xml_text)
        doc = _parse_fragments_to_dom(frag)
        if not doc:
            return None
        fs_entries = []
        for fs in doc.getElementsByTagName('filesystem'):
            def _gt(tag):
                n = fs.getElementsByTagName(tag)
                if not n:
                    return ''
                v = n[0]
                txt = v.firstChild.data.strip() if v.firstChild else ''
                if not txt and v.hasAttribute('junos:format'):
                    txt = v.getAttribute('junos:format')
                return txt
            total_txt = _gt('total-blocks')
            used_txt  = _gt('used-blocks')
            avail_txt = _gt('available-blocks')
            perc_txt  = _gt('used-percent')
            mnt_txt   = _gt('mounted-on')
            if not mnt_txt:
                mnt_txt = _gt('filesystem-name')
            total = _to_mb(total_txt)
            used  = _to_mb(used_txt)
            free  = _to_mb(avail_txt)
            try:
                util = int(re.sub('[^0-9]','', perc_txt)) if perc_txt else (int(round((used/total)*100)) if total>0 else 0)
            except Exception:
                util = int(round((used/total)*100)) if total>0 else 0
            fs_entries.append((mnt_txt, total, used, free, util))
        if not fs_entries:
            return None
        best_by_pref = None; best_fallback = None; max_total = -1
        for (mount, total, used_mb, free_mb, util) in fs_entries:
            for pat in PREF_MOUNT_PATTERNS:
                if re.search(pat, mount):
                    best_by_pref = (total, used_mb, free_mb, util, mount)
                    break
            if total > max_total:
                max_total = total; best_fallback = (total, used_mb, free_mb, util, mount)
        picked = best_by_pref or best_fallback
        return {
            'total_mb': int(round(picked[0])),
            'used_mb' : int(round(picked[1])),
            'free_mb' : int(round(picked[2])),
            'util_percent': int(round(picked[3])),
            'chosen': picked[4],
        }
    except Exception:
        return None

def _parse_storage_text(storage_text: str):
    best = None; best_by_pref = None; max_total = -1.0; chosen_line = ''
    for raw_line in str(storage_text or '').splitlines():
        line = raw_line.strip()
        if not line or ('Filesystem' in line and 'Mounted on' in line):
            continue
        if '%' not in line or '/' not in line:
            continue
        toks = re.findall(r'(\d+(?:\.\d+)?[KMG])', line)
        total = used = free = None
        if len(toks) >= 3:
            total = _to_mb(toks[0]); used = _to_mb(toks[1]); free = _to_mb(toks[2])
        else:
            nums = re.findall(r'\b(\d+(?:\.\d+)?)\b', line)
            if len(nums) >= 3:
                total = float(nums[0]); used = float(nums[1]); free = float(nums[2])
            else:
                continue
        mper = re.search(r'(\d+)%', line)
        util = float(mper.group(1)) if mper else (round((used/total)*100) if (total>0) else 0)
        for pat in PREF_MOUNT_PATTERNS:
            if re.search(pat, line):
                best_by_pref = (total, used, free, util); chosen_line = line
                break
        if total > max_total:
            max_total = total; best = (total, used, free, util); chosen_line = line
    picked = best_by_pref or best
    return {
        'total_mb': int(round(picked[0])) if picked else None,
        'used_mb' : int(round(picked[1])) if picked else None,
        'free_mb' : int(round(picked[2])) if picked else None,
        'util_percent': int(round(picked[3])) if picked else None,
        'chosen': chosen_line,
    }


# --- Temperature parser from RE (not CPU temperature)
def _parse_temperature_from_re(show_re_text: str):
    """Return temperature (°C) parsed from 'show chassis routing-engine' text.
    Picks lines containing 'Temperature' but ignores 'CPU temperature'."""
    maxc = None
    for line in str(show_re_text or '').splitlines():
        l = line.strip()
        if not l:
            continue
        if re.search(r'CPU\s+temperature', l, flags=re.I):
            continue
        if re.search(r'\bTemperature\b', l, flags=re.I):
            m = re.search(r'\b(\d{2,3})\b\s*(?:degrees\s*C|Celsius|C\b)?', l)
            if m:
                val = int(m.group(1))
                if 10 <= val <= 120:
                    maxc = val if maxc is None else max(maxc, val)
    return float(maxc) if maxc is not None else None
class JunosCliSession:
    _paging_disabled = False
    def __init__(self, channel):
        self.chan = channel
        try: self.chan.settimeout(5.0)
        except Exception: pass
        try: self.chan.set_combine_stderr(True)
        except Exception: pass
    def send(self, cmd: str):
        if not cmd.endswith('\n'): cmd += '\n'
        self.chan.send(cmd)
    def recv_until_prompt(self, prompt_regex=r'(>\s*$\n}\s*$\n%\s*$)', timeout=5):
        import re, time
        buf = ''; end = time.monotonic() + timeout
        while time.monotonic() < end:
            if self.chan.recv_ready():
                chunk = self.chan.recv(65536).decode('utf-8', errors='ignore'); buf += chunk
                if re.search(prompt_regex, buf, re.MULTILINE): break
            else: time.sleep(0.05)
        return buf
    def recv_until_tag_close(self, close_tag='</rpc-reply>', timeout=60, also_require_prompt=False, prompt_regex=r'(>\s*$\n}\s*$\n%\s*$)'):
        import re, time
        buf = ''; end = time.monotonic() + timeout
        while time.monotonic() < end:
            if self.chan.recv_ready():
                chunk = self.chan.recv(131072).decode('utf-8', errors='ignore'); buf += chunk
                if close_tag in buf:
                    if not also_require_prompt: break
                    if re.search(prompt_regex, buf, re.MULTILINE): break
            else: time.sleep(0.05)
        return buf




# Ensure router CLI has no paging (called once per node)
def ensure_no_paging(session):
    try:
        if getattr(session, '_paging_disabled', False):
            return
        session.send('set cli screen-length 0\n')
        _ = session.recv_until_prompt(timeout=3)
        _ = session.recv_until_prompt(timeout=3)
        session._paging_disabled = True
    except Exception:
        pass
def open_tacacs_shell(host, username, password, port=SSH_PORT):
    client = paramiko.SSHClient(); client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(hostname=host, username=username, password=password, port=port, look_for_keys=False, allow_agent=False, timeout=10, banner_timeout=BANNER_TIMEOUT, compress=True)
    try:
        client.get_transport().set_keepalive(30)
    except Exception:
        pass
    chan = client.invoke_shell(width=200, height=50); return client, JunosCliSession(chan)


def _feed_yes_no_if_needed(sess: JunosCliSession, buf: str, timeout=10):
    end = time.monotonic() + timeout; cur = buf or ''
    while time.monotonic() < end:
        low = cur.lower()
        if 'are you sure you want to continue connecting' in low or ('please type' in low and 'yes' in low and 'no' in low):
            sess.send('yes'); cur += sess.recv_until_prompt(timeout=3); break
        time.sleep(0.05); more = sess.recv_until_prompt(timeout=1)
        if more: cur += more
    return cur


def connect_to_node(sess: JunosCliSession, node: str, router_user: str = None, router_pass: str = None):
    if not router_user: return
    sess.send(f"ssh -o StrictHostKeyChecking=no -o UserKnownHostsFile=/dev/null {router_user}@{node}")
    out = sess.recv_until_prompt(timeout=10); out = _feed_yes_no_if_needed(sess, out, timeout=5)
    if 'password:' in out.lower(): sess.send(router_pass or ''); _ = sess.recv_until_prompt(timeout=10)


def recv_until_xml_or_prompt(session: JunosCliSession, timeout=60):
    import re, time
    buf = ''; end = time.monotonic() + timeout
    while time.monotonic() < end:
        if session.chan.recv_ready():
            chunk = session.chan.recv(65536).decode('utf-8', errors='ignore'); buf += chunk
            if '</rpc-reply>' in buf: break
            if re.search(r'(>\s*$\n}\s*$\n%\s*$)', buf, re.MULTILINE): break
        else: time.sleep(0.05)
    return buf


def junos_run_text(session, cmd, timeout=20):
    try:
        ensure_no_paging(session)
        session.send(cmd + "\n")
        out = session.recv_until_prompt(timeout=timeout)
        logger.debug(f"junos_run_text cmd='{cmd}' len={len(out)}")
        return out
    except Exception as e:
        append_error_log(get_debug_log_path('junos_text_errors.log'), f"Junos TEXT run failed for '{cmd}': {e}")
        return ''

# --- loopback parse ---
def parse_loopback0_ip(xml_or_text):
    if not xml_or_text: return '-'
    s = str(xml_or_text)
    try:
        frag = _extract_xml_fragment(s); doc = _parse_fragments_to_dom(frag)
        if doc:
            for li in doc.getElementsByTagName('logical-interface'):
                name = _get_first_text(li, 'name')
                if name and name.strip().lower() == 'lo0.0':
                    for af in li.getElementsByTagName('address-family'):
                        afn = _get_first_text(af, 'address-family-name')
                        if afn and 'inet' in afn.lower():
                            for ia in af.getElementsByTagName('interface-address'):
                                val = _get_first_text(ia, 'ifa-local')
                                if __is_ipv4(val or ''): return val
            for c in doc.getElementsByTagName('ifa-local') if doc else []:
                val = c.firstChild.data.strip() if c.firstChild else ''
                if __is_ipv4(val): return val
    except Exception:
        pass
    m = re.search(r'<ifa-local>\s*(\d{1,3}(?:\.\d{1,3}){3})\s*</ifa-local>', s)
    if m: return m.group(1)
    m2 = re.search(r'\binet\s+(\d{1,3}(?:\.\d{1,3}){3})\b', s, re.IGNORECASE)
    return m2.group(1) if m2 else '-'


def __is_ipv4(val):
    try:
        parts = str(val).split('.')
        return len(parts) == 4 and all(0 <= int(p) <= 255 for p in parts)
    except Exception: return False

# --- xml display/save helpers ---
def junos_xml(session: JunosCliSession, base_cmd: str) -> str:
    try:
        session.send("set cli screen-length 0"); _ = session.recv_until_prompt(timeout=3)
        ensure_no_paging(session)
        session.send(f"{base_cmd} | display xml\n")
        out = recv_until_xml_or_prompt(session, timeout=60)
        logger.debug(f"junos_xml cmd='{base_cmd}' len={len(out)}")
        return out
    except Exception as e:
        append_error_log(get_debug_log_path('junos_xml_errors.log'), f"Junos XML run failed for '{base_cmd}': {e}")
        return ''


def _send_and_expect(session: JunosCliSession, cmd: str, expect_regex=r'(>\s*$\n}\s*$\n%\s*$\nsave complete)', timeout=10) -> str:
    session.send(cmd); out = session.recv_until_prompt(expect_regex, timeout=timeout); logger.debug(f"_send_and_expect cmd='{cmd}' len={len(out)}"); return out


def _safe_xml_filename(node: str, tag: str) -> str:
    base_node = re.sub(r'[^A-Za-z0-9_.\-]+', '_', str(node or 'node'))
    base_tag = re.sub(r'[^A-Za-z0-9_.\-]+', '_', str(tag or 'xml'))
    return f"/var/tmp/{base_node}_{base_tag}.xml"


def junos_xml_save_and_read(session: JunosCliSession, node_name: str, base_cmd: str, tag_hint: str) -> str:
    try:
        try: session.send("set cli screen-length 0"); _ = session.recv_until_prompt(timeout=2)
        except Exception: pass
        tmp_path = _safe_xml_filename(node_name, tag_hint)
        _ = _send_and_expect(session, f"{base_cmd}\n display xml\n save {tmp_path}", timeout=30)
        session.send(f"file show {tmp_path}\n")
        out_show = session.recv_until_tag_close(close_tag='</rpc-reply>', timeout=60, also_require_prompt=False)
        if '</rpc-reply>' not in out_show:
            try:
                lst = _send_and_expect(session, f"file list {tmp_path} detail", timeout=10)
                m_sz = re.search(r'(?i)\bsize\s*[:=]?\s*(\d{3,})\b', lst) or re.search(r'(?i)\b(\d{3,})\s*bytes\b', lst)
                sz = int(m_sz.group(1)) if m_sz else 0
            except Exception:
                sz = 0
            session.send(f"file show {tmp_path}\n"); buf = ''; end = time.monotonic() + 60
            while time.monotonic() < end:
                if session.chan.recv_ready():
                    chunk = session.chan.recv(131072).decode('utf-8', errors='ignore'); buf += chunk
                    if '</rpc-reply>' in buf: break
                    pay = sanitize_xml_text(buf)
                    if sz and len(pay.encode('utf-8', errors='ignore')) >= sz: break
                else: time.sleep(0.05)
            out_show = buf
        xml_text = sanitize_xml_text(out_show); logger.debug(f"junos_xml_save_and_read base_cmd='{base_cmd}' size={len(xml_text)} path={tmp_path}"); return xml_text or out_show
    except Exception as e:
        append_error_log(get_debug_log_path('junos_xml_errors.log'), f"OpsiA save/read failed for '{base_cmd}' ({node_name}): {e}")
        return ''

# ---------- Parallel collector ----------
from concurrent.futures import ThreadPoolExecutor, as_completed

def _collect_for_node(node: str, tacacs_chosen: str, tacacs_user: str, tacacs_pass: str, router_user: str, router_pass: str) -> Dict[str, Any]:
    t0 = time.monotonic()
    out: Dict[str, Any] = {
        'node': node,
        'hardware_items': [],
        'interfaces_rows': [],
        'alarms': [],
        'loopback': '-',
        'system_info': {
            'platform':'mx960','current_sw':'JUNOS', 'loopback_address':'-',
            'memory_util':3,'cpu_usage':5,
            'total_space':54272,'used_space':6144,'free_space':48128,'disk_util':3,'temperature':49,
            'memory_recommendation':'NORMAL - Optimal Performance',
            'cpu_recommendation':'NORMAL - Optimal Performance',
            'disk_recommendation':'NORMAL - Adequate Free Space',
        },
        'fpc_model_map': {},
        'optics_map': {},
        'errors': []
    }
    node_log = get_debug_log_path(f"NODE_{node}_{capture_time_global.strftime('%H%M%S')}.log")
    fh = logging.FileHandler(node_log, mode='w', encoding='utf-8')
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter('%(asctime)s %(levelname)s - %(message)s'))
    logger.addHandler(fh)
    print_status('DEBUG', f'Start collect for node {node}', node)
    client = None
    try:
        client, sess = open_tacacs_shell(tacacs_chosen, tacacs_user, tacacs_pass, port=SSH_PORT)
        try:
            connect_to_node(sess, node, router_user=router_user, router_pass=router_pass)
        except Exception as e:
            out['errors'].append(f'connect_to_node: {e}')
            print_status('WARN', f'connect_to_node error: {e}', node)
        # Temperature from RE (not CPU temperature) + CPU/Memory/Disk collectors
        try:
            re_txt = junos_run_text(sess, 'show chassis routing-engine | no-more', timeout=40)
            save_log(get_debug_log_path(f"{node}_routing-engine.txt"), re_txt)
            # Temperature
            temp_c = _parse_temperature_from_re(re_txt)
            if temp_c is not None:
                out['system_info']['temperature'] = int(temp_c)
            # CPU & Memory
            mem_txt = junos_run_text(sess, 'show system memory', timeout=40)
            save_log(get_debug_log_path(f"{node}_system-memory.txt"), mem_txt)
            cpu_used = _parse_cpu_used_from_re(re_txt)
            mem_util = _parse_mem_util_from_re_or_sysmem(re_txt, mem_txt)
            if cpu_used is not None:
                out['system_info']['cpu_usage'] = int(cpu_used)
            if mem_util is not None:
                out['system_info']['memory_util'] = int(mem_util)
            print_status('DEBUG', f"temp(RE)={out['system_info'].get('temperature',0)}C cpu={out['system_info'].get('cpu_usage',0)}% mem={out['system_info'].get('memory_util',0)}%", node)
        except Exception as e:
            out['errors'].append(f're/cpu/mem: {e}')
        # Storage XML prefer → text fallback
        try:
            stg_xml = junos_xml(sess, 'show system storage')
            save_log(get_debug_log_path(f"{node}_system-storage.xml"), stg_xml)
        except Exception:
            stg_xml = ''
        try:
            stg_txt = junos_run_text(sess, 'show system storage', timeout=55)
            save_log(get_debug_log_path(f"{node}_system-storage.txt"), stg_txt)
        except Exception:
            stg_txt = ''
        st_xml = _parse_storage_xml(stg_xml)
        st_txt = _parse_storage_text(stg_txt)
        chosen = st_xml if st_xml and all(st_xml.get(k) is not None for k in ('total_mb','used_mb','free_mb','util_percent')) else st_txt
        try:
            out['system_info']['total_space'] = int((chosen or {}).get('total_mb') or 0)
            out['system_info']['used_space']  = int((chosen or {}).get('used_mb')  or 0)
            out['system_info']['free_space']  = int((chosen or {}).get('free_mb')  or 0)
            out['system_info']['disk_util']   = int((chosen or {}).get('util_percent') or (
                round((((chosen or {}).get('used_mb') or 0) / max(((chosen or {}).get('total_mb') or 1), 1)) * 100)
            ))
            print_status('DEBUG', f"storage: total={out['system_info']['total_space']}MB used={out['system_info']['used_space']}MB free={out['system_info']['free_space']}MB util={out['system_info']['disk_util']}%", node)
        except Exception as e:
            out['errors'].append(f'storage-parse: {e}')

        try:
            xml_chassis = junos_xml(sess, 'show chassis hardware detail')
            save_log(get_debug_log_path(f"{node}_chassis.xml"), xml_chassis)
            out['hardware_items'] = parse_hardware_inventory(xml_chassis)
            print_status('DEBUG', f'hardware_items={len(out["hardware_items"]) }', node)
        except Exception as e:
            out['errors'].append(f'chassis: {e}')
        try:
            xml_optics = junos_xml(sess, 'show interfaces diagnostics optics')
            save_log(get_debug_log_path(f"{node}_optics.xml"), xml_optics)
            frag_opt = _extract_xml_fragment(xml_optics); doc_opt = _parse_fragments_to_dom(frag_opt)
            if doc_opt:
                for phys in doc_opt.getElementsByTagName('optics-diagnostics'):
                    parent = phys.parentNode
                    name = _get_first_text(parent,'name')
                    desc = (_get_first_text(phys,'module-type') or _get_first_text(phys,'module-description') or _get_first_text(phys,'vendor-name') or _get_first_text(phys,'model-number') or _get_first_text(phys,'part-number'))
                    if name and desc:
                        out['optics_map'][name]=desc; out['optics_map'][name.split('.')[0]]=desc
            print_status('DEBUG', f'optics_map={len(out["optics_map"]) }', node)
        except Exception as e:
            out['errors'].append(f'optics: {e}')
        try:
            xml_alarms = junos_xml(sess, 'show chassis alarms')
            save_log(get_debug_log_path(f"{node}_alarms.xml"), xml_alarms)
            out['alarms'] = parse_chassis_alarms(xml_alarms)
            print_status('DEBUG', f'alarms={len(out["alarms"]) }', node)
        except Exception as e:
            out['errors'].append(f'alarms: {e}')
        try:
            xml_ifaces = junos_xml(sess, 'show interfaces extensive')
            save_log(get_debug_log_path(f"{node}_interfaces.xml"), xml_ifaces)
            out['interfaces_rows'] = parse_interfaces_xml_basic(xml_ifaces)
            print_status('DEBUG', f'interfaces_rows={len(out["interfaces_rows"]) }', node)
        except Exception as e:
            out['errors'].append(f'interfaces: {e}')

        # Detect platform & software version (outside interface try); use RAW token in current_sw
        try:
            show_ver = junos_run_text(sess, 'show version', timeout=30)
        except Exception:
            show_ver = ''
        try:
            plat, os_type, raw_ver = _detect_platform_and_sw(node, out.get('hardware_items', []), show_ver)
            out['system_info']['platform'] = plat
            out['system_info']['current_sw'] = raw_ver
            out['system_info']['current_sw_type'] = os_type
            out['system_info']['show_version_text'] = show_ver
            print_status('DEBUG', f"Platform detected: {plat}", node)
            print_status('DEBUG', f"OS type: {os_type}, version: {raw_ver}", node)
        except Exception as e:
            out['errors'].append(f'platform/sw detect: {e}')
        try:
            xml_lo = junos_xml(sess, 'show interfaces lo0.0')
            out['loopback'] = parse_loopback0_ip(xml_lo)
            if out['loopback'] == '-':
                txt_lo = junos_run_text(sess, 'show interfaces lo0.0 terse', timeout=20)
                out['loopback'] = parse_loopback0_ip(txt_lo)
            print_status('DEBUG', f'loopback={out["loopback"]}', node)
        except Exception as e:
            out['errors'].append(f'loopback: {e}')
        try:
            out['fpc_model_map'] = build_fpc_model_map(out['hardware_items'])
            print_status('DEBUG', f'fpc_model_map={len(out["fpc_model_map"]) }', node)
        except Exception as e:
            out['errors'].append(f'fpc_map: {e}')
    except Exception as e:
        out['errors'].append(f'open_tacacs_shell: {e}')
        print_status('ERROR', f'open_tacacs_shell error: {e}', node)
    finally:
        try:
            if client: client.close()
        except Exception: pass
        try:
            logger.removeHandler(fh)
            fh.close()
        except Exception:
            pass
    out['elapsed'] = time.monotonic() - t0
    print_status('DEBUG', f'Finish collect in {out["elapsed"]:.1f}s', node)
    return out

# ---------- sheet summaries ----------
def add_all_sheet_summaries(wb, nodes):
    tz = get_indonesia_timezone()
    ws_main = wb[MAIN_SHEET]; total_ifaces = _count_data_rows(ws_main, start_row=6, must_have_cols=(2,))
    add_sheet_footer_summary(ws_main, "FPC UTILIZATION ANALYSIS SUMMARY",
                             [f"Total Interfaces Analyzed: {total_ifaces}", f"Analysis Date: {capture_time_global.strftime('%d %B %Y at %H:%M')} {tz}", f"Network Infrastructure Monitoring Report - Generated on {capture_time_global.strftime('%d %B %Y at %H:%M')} {tz}"])
    ws_util = wb[UTIL_SHEET]; total_ports = _count_data_rows(ws_util, start_row=6, must_have_cols=(2,))
    add_sheet_footer_summary(ws_util, "PORT UTILIZATION DETAILED ANALYSIS",
                             [f"Total Ports Analyzed: {total_ports}", f"Network Infrastructure Monitoring Report - Generated on {capture_time_global.strftime('%d %B %Y at %H:%M')} {tz}"])
    ws_alarm = wb[ALARM_SHEET]; total_alarm = _count_data_rows(ws_alarm, start_row=6, must_have_cols=(2,))
    add_sheet_footer_summary(ws_alarm, "NETWORK ALARM STATUS SUMMARY",
                             [f"Total Alarm Records: {total_alarm}", f"Network Infrastructure Monitoring Report - Generated on {capture_time_global.strftime('%d %B %Y at %H:%M')} {tz}"])
    ws_hw = wb[HARDWARE_SHEET]; total_hw = _count_data_rows(ws_hw, start_row=6, must_have_cols=(2,))
    add_sheet_footer_summary(ws_hw, "HARDWARE INVENTORY ANALYSIS",
                             [f"Total Hardware Components: {total_hw}", f"Network Infrastructure Monitoring Report - Generated on {capture_time_global.strftime('%d %B %Y at %H:%M')} {tz}"])
    ws_sys = wb[SYSTEM_SHEET]
    add_sheet_footer_summary(ws_sys, "SYSTEM PERFORMANCE MONITORING SUMMARY",
                             [f"Total Network Nodes Monitored: {len(nodes)}", f"Network Infrastructure Monitoring Report - Generated on {capture_time_global.strftime('%d %B %Y at %H:%M')} {tz}"])
    logger.debug("add_all_sheet_summaries completed")

# ---------- main ----------
def main():
    global folder_daily_global, folder_monthly_global, capture_time_global, debug_folder_global, _progress_start
    desktop = get_desktop_path(); capture_time_global = datetime.datetime.now()
    folder_monthly_global = os.path.join(desktop, 'LAB-Occupancy'); os.makedirs(folder_monthly_global, exist_ok=True)
    folder_daily_global = os.path.join(folder_monthly_global, 'Capture_FPC-Occupancy' + capture_time_global.strftime('%Y%m%d')); os.makedirs(folder_daily_global, exist_ok=True)
    debug_folder_global = os.path.join(folder_daily_global, 'All Debug'); os.makedirs(debug_folder_global, exist_ok=True); setup_debug_folder()
    run_log_file = get_debug_log_path(f"RUN_{capture_time_global.strftime('%Y%m%d_%H%M%S')}.log")
    logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(levelname)s - %(message)s', handlers=[logging.FileHandler(run_log_file, encoding='utf-8'), logging.StreamHandler(sys.stdout)])

    access_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'access_lab.xml')
    node_list_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'list_lab.txt')
    excel_file = os.path.join(folder_monthly_global, 'LAB_Occupancy_Report_' + capture_time_global.strftime('%d%b%Y_%H%M') + '.xlsx')

    if not os.path.exists(access_file): sys.stderr.write(f'Access file not found: {access_file}\n'); sys.exit(1)
    try: access_doc = minidom.parse(access_file)
    except Exception as e: sys.stderr.write(f'Failed to parse access file: {e}\n'); sys.exit(1)

    tacacs_user = ''; tacacs_pass = ''; router_pass = ''
    try:
        tacacs_user = access_doc.getElementsByTagName('tacacs-user')[0].firstChild.data
        tacacs_pass = access_doc.getElementsByTagName('tacacs-pass')[0].firstChild.data
    except Exception: pass
    try: router_pass = access_doc.getElementsByTagName('router-pass')[0].firstChild.data
    except Exception: router_pass = ''

    tacacs_list = access_doc.getElementsByTagName('tacacs-server'); tacacs_chosen = None
    for i in range(tacacs_list.length):
        tac = tacacs_list[i].firstChild.data; tried = 0
        client = paramiko.SSHClient(); client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        while tried <= INITIAL_TEST_RETRIES:
            tried += 1
            try:
                client.connect(hostname=tac, username=tacacs_user, password=tacacs_pass, port=SSH_PORT, look_for_keys=False, allow_agent=False, timeout=10, banner_timeout=BANNER_TIMEOUT)
                client.close(); tacacs_chosen = tac; break
            except Exception as e:
                append_error_log(os.path.join(folder_daily_global, f'_KONEKSI_{tac}_GAGAL_{capture_time_global.strftime("%Y%m%d_%H%M")}.log'), f'Initial connect failed to {tac}: {e}')
                if tried <= INITIAL_TEST_RETRIES: time.sleep(INITIAL_TEST_RETRY_DELAY)
        try: client.close()
        except Exception: pass
        if tacacs_chosen: break
    if not tacacs_chosen: sys.stderr.write('Failed to connect to any TACACS server. See logs.\n'); sys.exit(2)

    if not os.path.exists(excel_file): workbook_create(excel_file)
    try: wb = load_workbook(excel_file)
    except Exception as e: sys.stderr.write(f'Failed to load workbook: {e}\n'); sys.exit(1)

    if not os.path.exists(node_list_file): sys.stderr.write(f'Node list not found: {node_list_file}\n'); sys.exit(1)
    with open(node_list_file, 'r', encoding='utf-8', errors='ignore') as f: nodes = [ln.strip() for ln in f if ln.strip()]

    _banner_start(len(nodes))
    _progress_start = time.monotonic()

    results_per_node: List[Dict[str, Any]] = []
    try: router_user = access_doc.getElementsByTagName('router-user')[0].firstChild.data
    except Exception: router_user = None

    print_status('INFO', f'Using TACACS: {tacacs_chosen}')
    # SAFE parallelism: env override or CPU-based with hard cap
    max_workers_env = os.getenv('TELKOM_MAX_WORKERS')
    if max_workers_env and max_workers_env.isdigit():
        max_workers = max(1, int(max_workers_env))
    else:
        io_factor = int(os.getenv('TELKOM_IO_FACTOR', '4'))
        hard_cap  = int(os.getenv('TELKOM_MAX_CAP', '64'))
        max_workers = min(len(nodes), (os.cpu_count() or 4) * io_factor, hard_cap)
    print_status('INFO', f'Parallel workers: {max_workers}')

    t_pool_start = time.monotonic()
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        futures = {ex.submit(_collect_for_node, node, tacacs_chosen, tacacs_user, tacacs_pass, router_user, router_pass): node for node in nodes}
        for idx, fut in enumerate(as_completed(futures), 1):
            node = futures[fut]
            print_progress(idx, len(nodes), node, operation='Collecting')
            try:
                res = fut.result()
                results_per_node.append(res)
                print_status('INFO', f'Node collected in {res.get("elapsed",0):.1f}s', node)
            except Exception as e:
                print_status('ERROR', f'Thread failed: {e}', node)
    print_status('INFO', f'Parallel collection completed in {(time.monotonic()-t_pool_start):.1f}s')

    # --- Write to Excel sequentially ---
    system_results = {}
    for res in results_per_node:
        node = res['node']
        for it in res.get('hardware_items', []):
            write_hardware_row_simple(node, get_area_pop_from_node(node), it.get('component_type'), it.get('slot'), it.get('part'), it.get('serial'), it.get('model'), it.get('version'), it.get('status'), it.get('comments'), wb)
        fpc_model_map = res.get('fpc_model_map', {})
        optics_map = res.get('optics_map', {})
        rows = [r for r in res.get('interfaces_rows', []) if str(r.get('iface','')).startswith(('ae-','et-','xe-','ge-'))]
        for r in rows:
            iface = r.get('iface',''); desc = r.get('desc',''); cap = r.get('capacity',''); util = float(r.get('util',0.0)); gb = float(r.get('traffic_gb',0.0))
            last_flapped = r.get('last_flapped','')
            module_type = optics_map.get(iface, optics_map.get(iface.split('.') [0], None))
            m_fpc = re.match(r'^(?:et|xe|ge)-(\d+)/', iface)
            if m_fpc:
                fpc_slot = int(m_fpc.group(1)); hw_model = fpc_model_map.get(fpc_slot)
                if hw_model: module_type = hw_model
            if not module_type: module_type = 'Aggregated Ethernet Bundle' if iface.startswith('ae-') else 'Ethernet'
            write_data_row_simple(node, get_area_pop_from_node(node), desc, iface, module_type, cap, gb, util, _status_color(util), wb)
            port_status = 'USED' if util > 0 else 'UNUSED'; configured = 'Yes' if port_status == 'USED' else 'No'
            opt_info = optics_map.get(iface, optics_map.get(iface.split('.') [0], optics_map.get(iface.split('-')[0], '')))
            sfp_status = str(opt_info) if opt_info else ('QSFP Module' if iface.startswith('et-') else 'SFP+ Module' if iface.startswith('xe-') else 'SFP Module' if iface.startswith('ge-') else 'Unknown')
            write_utilisasi_port_row_simple(node, get_area_pop_from_node(node), iface, module_type, cap, last_flapped, sfp_status, configured, desc, port_status, 'Stable', wb)
        alarms = res.get('alarms', [])
        if not alarms:
            write_alarm_row_simple(node, get_area_pop_from_node(node), None, None, None, None, None, wb)
        else:
            for a in alarms:
                write_alarm_row_simple(node, get_area_pop_from_node(node), a.get('time'), a.get('type'), a.get('description'), a.get('severity'), a.get('status'), wb)
        sys_info = res.get('system_info', {})
        # Safety net: ensure platform/SW filled if collector failed (use RAW token in current_sw)
        try:
            show_ver_text = res.get('system_info', {}).get('show_version_text', '')
            plat, os_type, raw_ver = _detect_platform_and_sw(node, res.get('hardware_items', []), show_ver_text)
            sys_info['platform'] = plat
            sys_info['current_sw'] = raw_ver
            sys_info['current_sw_type'] = os_type
        except Exception as e:
            append_error_log(get_debug_log_path('platform_sw_detect.log'), f"{node}: {e}")

        sys_info['loopback_address'] = res.get('loopback','-')
        system_results[node] = sys_info

    try:
        worksheet_system_performance(wb[SYSTEM_SHEET], system_data=system_results)
        finalize_tables(wb); apply_conditional_formatting(wb)
        add_all_sheet_summaries(wb, nodes)
        populate_dashboard_like_example(wb, nodes)
        ws_main = wb[MAIN_SHEET]
        for r in range(6, ws_main.max_row+1): ws_main[f'A{r}'].number_format = '0'
        wb.save(excel_file); wb.close()
        total_secs = time.monotonic() - _progress_start
        _banner_end(total_secs)
        print(f"[OK] Report saved: {excel_file}")
    except Exception as e:
        print_status('ERROR', f'Finalize/Save failed: {e}')
        append_error_log(os.path.join(folder_daily_global, 'excel_save_errors.log'), f'Finalize/Save: {e}')

if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument('--ssh-port', type=int, default=None, help='Port SSH TACACS/jump-host')
    args = ap.parse_args()
    if args.ssh_port:
        SSH_PORT = args.ssh_port
    main()
